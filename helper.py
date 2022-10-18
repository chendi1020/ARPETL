from nbformat import from_dict
import yaml
import os
import numpy as np 
import pandas as pd 
from openpyxl import load_workbook
import logging
from pathlib import Path
import inspect
def import_config(path):
    """
    Imports config file into a python dictionary
    Args:
        path (Path obj): file path for config file
    Returns:
        config dict (dict): dict for config items
    """
    with open(path) as f :
        return yaml.safe_load(f)


def output_excel(df, path, filename):
    df.to_excel(os.path.join(path, filename), index=False)


def read_link(file, sheet, starti, endi, col):
    """
    return df with rowid and url
    """
    wb = load_workbook(file)
    ws = wb.get_sheet_by_name(sheet)
    hylink={}
    for i in range(starti,endi):
        s=ws.cell(row=i, column=ord(col.lower())-96).hyperlink
        if s==None:
            hylink[i]= None
        else:
            hylink[i]=s.target
    hylinkdf =pd.DataFrame(list(hylink.items()),columns = ['rowid','link']) 
    return hylinkdf


us_state_to_abbrev = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "District of Columbia": "DC",
    "Washington, DC":"DC",
    "American Samoa": "AS",
    "Guam": "GU",
    "Northern Mariana Islands": "MP",
    "Puerto Rico": "PR",
    "United States Minor Outlying Islands": "UM",
    "U.S. Virgin Islands": "VI",
}
    
# invert the dictionary
#abbrev_to_us_state = dict(map(reversed, us_state_to_abbrev.items()))
STat = pd.Series(us_state_to_abbrev)
STat.name= 'STAbbr'

def sort_jurisidiction(df):
    df = df.sort_values(['Level_of_Goverment','State', 'Jurisdication'])
    df=df.join(STat, on='State')
    #recode jursdiction 
    df['Jurisdication']= np.where(df['Level_of_Goverment'] !='County', df['Jurisdication'],
     np.where(df['Jurisdication'].str.contains('County'), df['Jurisdication'], df['Jurisdication'].str.strip()+ " County" ))
    df['Jurisdiction']= np.where(df['Level_of_Goverment']=='State', df['STAbbr'],
        np.where(df['STAbbr']=="DC", "District of Columbia",  df['Jurisdication']+","+df['STAbbr'] )
      )
    return df 

def check_STAbbr(df):
    if df['STAbbr'].isnull().sum()>0:
        t=df[df['STAbbr'].isnull()].loc[:,['State','Jurisdication']]
        return t
    else:
        return("no issue")


InvestAreaECMap={'Behavioral health':'EC1',
 'Broadband': 'EC5',
 'Building affordable housing':'EC2',
 'Building housing for people experiencing homelessness':'EC2',
 'Career prep programs':'EC2',
 'Childcare':'EC2',
 'City/county - school district collaboration':"NA",
 'City/county collaboration':"NA",
 'Community development':"NA",
 'Early childhood programs':'EC2',
 'Education':'EC2',
 'Educational tech/broadband access':'EC5',
 'Financial stablity':'EC2',
 'Food insecurity (including snap benefits)':'EC2',
 'Guaranteed income/cash transfers':'EC2',
 'High-dosage tutoring programs':'EC2',
 'Home/rental assistance':'EC2',
 'Infant/childcare and maternal health':'EC2',
 'Justice':"NA",
 'Local business support':'EC2',
 'Mental health':'EC1',
 'Older adults':'EC2',
 'Other':"NA",
 'Public health':'EC1',
 'Public safety':'EC2',
 'School facility improvements':'EC2',
 'Services for people experiencing homelessness':'EC2',
 'State/local government collaboration':"NA",
 'Substance use':'EC1',
 'Summer/extended learning programs':'EC2',
 'Workforce development & training':'EC3',
}


#function merge with check
def mergechk(leftdf, rdf, mergebycol,checkcol):
    t = leftdf.join(rdf.set_index(mergebycol), on= mergebycol)
    if len(t) != len(leftdf):
        print('records does not match')
    if t[checkcol].isnull().sum()>0:
        print(t[t[checkcol].isnull()].loc[:,mergebycol].drop_duplicates())
    else:
        print("all matched")


#basic logging
def setupLog():
    
    #create custome logger
    caller_path = Path(inspect.stack()[1][1])
    logger = logging.getLogger(caller_path.name)
    #create handler
    c_handler = logging.StreamHandler()
    #create format and add to handler
    c_format = logging.Formatter('%(name)s - %(levelname)s -%(message)s')
    c_handler.setFormatter(c_format)
    #set level
    logger.setLevel(logging.INFO)
    #add handler to the logger
    logger.addHandler(c_handler)
    return logger

def output_to_excel(path, df,  outputname):
    """
    path of the output
    outputname: name of the output file
    """
    df.to_excel(os.path.join(path, outputname), index=False)